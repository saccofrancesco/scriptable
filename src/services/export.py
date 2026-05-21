from __future__ import annotations
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from ..domain.models import Employee, Schedule, Shift
from .calculations import month_day_count
from .formatting import format_month_label


@dataclass(frozen=True, slots=True)
class ShiftExportRow:
    employee_id: str
    employee_name: str
    shift_date: date
    start_time: time
    end_time: time
    duration_hours: float
    notes: str = ""


def _solid_fill(color_hex: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color_hex.replace("#", "").upper())


def _apply_border(cell) -> None:
    thin: Side = Side(style="thin", color="D7DCE3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header_row(sheet, row_index: int, columns: int) -> None:
    fill: PatternFill = PatternFill(fill_type="solid", fgColor="EAF0FF")
    for column in range(1, columns + 1):
        cell = sheet.cell(row=row_index, column=column)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_border(cell)


def _set_page_setup(sheet) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    sheet.sheet_view.showGridLines = False


def _set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _freeze_after_headers(sheet, row: int = 1, column: int = 1) -> None:
    sheet.freeze_panes = f"{get_column_letter(column + 1)}{row + 1}"


def _month_days(month_date: date) -> list[date]:
    return [month_date.replace(day=day) for day in range(1, month_day_count(month_date) + 1)]


def _shift_duration_hours(shift: Shift) -> float:
    start: datetime = datetime.combine(shift.shift_date, shift.start_time)
    end: datetime = datetime.combine(shift.shift_date, shift.end_time)
    return max((end - start).total_seconds() / 3600, 0.0)


def _build_export_rows(schedule: Schedule, month_date: date) -> list[ShiftExportRow]:
    employees_by_id: dict[str, Employee] = {employee.id: employee for employee in schedule.employees}
    month_key: tuple[int, int] = (month_date.year, month_date.month)
    rows: list[ShiftExportRow] = []
    for shift in sorted(
        (
            shift
            for shift in schedule.shifts
            if (shift.shift_date.year, shift.shift_date.month) == month_key
        ),
        key=lambda shift: (
            shift.shift_date,
            shift.start_time,
            shift.end_time,
            employees_by_id.get(shift.employee_id).full_name.casefold()
            if employees_by_id.get(shift.employee_id)
            else "unknown",
        ),
    ):
        employee: Employee | None = employees_by_id.get(shift.employee_id)
        rows.append(
            ShiftExportRow(
                employee_id=shift.employee_id,
                employee_name=employee.full_name if employee else "Unknown employee",
                shift_date=shift.shift_date,
                start_time=shift.start_time,
                end_time=shift.end_time,
                duration_hours=_shift_duration_hours(shift),
            )
        )
    return rows


def _write_shifts_data_sheet(sheet, rows: list[ShiftExportRow], month_date: date) -> None:
    headers: list[str] = [
        "Employee",
        "Date",
        "StartTime",
        "EndTime",
        "Duration",
        "Notes",
    ]
    sheet.title = "Shifts_Data"
    title_row: int = 1
    subtitle_row: int = 2
    header_row: int = 4
    sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(headers))
    sheet.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=len(headers))
    sheet.cell(row=title_row, column=1, value=f"Shifts Data - {format_month_label(month_date)}").font = Font(
        size=15, bold=True, color="0F172A"
    )
    sheet.cell(
        row=subtitle_row,
        column=1,
        value="Source of truth for shift assignments",
    ).font = Font(size=10, color="475569")
    sheet.cell(row=title_row, column=1).alignment = Alignment(horizontal="left")
    sheet.cell(row=subtitle_row, column=1).alignment = Alignment(horizontal="left")

    for index, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=index, value=header)
    _style_header_row(sheet, header_row, len(headers))

    start_row: int = header_row + 1
    for index, row in enumerate(rows, start=start_row):
        employee_cell = sheet.cell(row=index, column=1, value=row.employee_name)
        date_cell = sheet.cell(row=index, column=2, value=row.shift_date)
        start_cell = sheet.cell(row=index, column=3, value=row.start_time)
        end_cell = sheet.cell(row=index, column=4, value=row.end_time)
        duration_cell = sheet.cell(row=index, column=5, value=row.duration_hours)
        notes_cell = sheet.cell(row=index, column=6, value=row.notes or "")

        date_cell.number_format = "yyyy-mm-dd"
        start_cell.number_format = "hh:mm"
        end_cell.number_format = "hh:mm"
        duration_cell.number_format = "0.00"

        for cell in (employee_cell, date_cell, start_cell, end_cell, duration_cell, notes_cell):
            cell.alignment = Alignment(vertical="center")
            _apply_border(cell)

    _freeze_after_headers(sheet, header_row, 1)
    sheet.auto_filter.ref = f"A{header_row}:F{max(start_row + len(rows) - 1, header_row)}"
    _set_page_setup(sheet)
    _set_widths(sheet, {"A": 26, "B": 14, "C": 12, "D": 12, "E": 12, "F": 24})


def _write_monthly_view_sheet(sheet, schedule: Schedule, month_date: date, rows: list[ShiftExportRow]) -> None:
    employees: list[Employee] = list(schedule.employees)
    days: list[date] = _month_days(month_date)
    sheet.title = "Monthly_View"

    headers: list[str] = ["Employee"] + [str(day.day) for day in days]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header)
    _style_header_row(sheet, 1, len(headers))

    shifts_by_employee_day: dict[tuple[str, date], list[ShiftExportRow]] = defaultdict(list)
    for row in rows:
        shifts_by_employee_day[(row.employee_id, row.shift_date)].append(row)

    for row_index, employee in enumerate(employees, start=2):
        name_cell = sheet.cell(row=row_index, column=1, value=employee.full_name)
        name_cell.fill = _solid_fill(employee.color_hex)
        name_cell.font = Font(bold=True, color="0F172A")
        name_cell.alignment = Alignment(vertical="center")
        _apply_border(name_cell)

        for col_offset, day in enumerate(days, start=2):
            cell = sheet.cell(row=row_index, column=col_offset)
            values = shifts_by_employee_day.get((employee.id, day), [])
            if values:
                cell.value = "\n".join(
                    f"{shift.start_time.strftime('%H:%M')}–{shift.end_time.strftime('%H:%M')}"
                    for shift in values
                )
            else:
                cell.value = ""
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            _apply_border(cell)

    _freeze_after_headers(sheet, 1, 1)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(employees) + 1, 1)}"
    _set_page_setup(sheet)
    _set_widths(sheet, {"A": 26, **{get_column_letter(index): 12 for index in range(2, len(headers) + 1)}})


def _write_summary_sheet(sheet, schedule: Schedule, month_date: date, rows: list[ShiftExportRow]) -> None:
    sheet.title = "Summary"
    headers: list[str] = ["Employee", "Total Hours", "Working Days"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header)
    _style_header_row(sheet, 1, len(headers))

    totals_by_employee: dict[str, float] = defaultdict(float)
    working_days_by_employee: dict[str, set[date]] = defaultdict(set)
    for row in rows:
        totals_by_employee[row.employee_id] += row.duration_hours
        working_days_by_employee[row.employee_id].add(row.shift_date)

    employees: list[Employee] = list(schedule.employees)
    for row_index, employee in enumerate(employees, start=2):
        employee_cell = sheet.cell(row=row_index, column=1, value=employee.full_name)
        total_cell = sheet.cell(row=row_index, column=2, value=totals_by_employee.get(employee.id, 0.0))
        days_cell = sheet.cell(row=row_index, column=3, value=len(working_days_by_employee.get(employee.id, set())))

        employee_cell.fill = _solid_fill(employee.color_hex)
        employee_cell.font = Font(bold=True, color="0F172A")
        total_cell.number_format = "0.00"

        for cell in (employee_cell, total_cell, days_cell):
            cell.alignment = Alignment(vertical="center")
            _apply_border(cell)

    _freeze_after_headers(sheet, 1, 1)
    sheet.auto_filter.ref = f"A1:C{max(len(employees) + 1, 1)}"
    _set_page_setup(sheet)
    _set_widths(sheet, {"A": 26, "B": 14, "C": 14})


def export_schedule_xlsx(schedule: Schedule, month_date: date, file_path: str | Path) -> Path:
    path: Path = Path(file_path)
    rows: list[ShiftExportRow] = _build_export_rows(schedule, month_date)

    workbook: Workbook = Workbook()
    workbook.remove(workbook.active)

    shifts_sheet = workbook.create_sheet("Shifts_Data")
    _write_shifts_data_sheet(shifts_sheet, rows, month_date)

    monthly_view_sheet = workbook.create_sheet("Monthly_View")
    _write_monthly_view_sheet(monthly_view_sheet, schedule, month_date, rows)

    summary_sheet = workbook.create_sheet("Summary")
    _write_summary_sheet(summary_sheet, schedule, month_date, rows)

    workbook.save(path)
    return path
